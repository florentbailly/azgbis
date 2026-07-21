"""Export Excel des transactions DVF de la zone de contexte (spec §7).

Mêmes jointures et même périmètre que le thème Marché de l'analyse (zone de
contexte = grand rayon) : le fichier téléchargé ne peut pas diverger de ce que
l'expert a vu à l'écran. Une ligne par local vendu (granularité DVF native).
"""
import asyncio
import io
import json

from . import db
from .geo import resolve_zone
from .schemas import ZoneInput

PLAFOND_LIGNES = 10_000  # garde-fou : une zone très large ne doit pas figer l'API

SQL = f"""
SELECT m.date_mutation, m.nature_mutation, m.valeur_fonciere, m.code_commune,
       l.id_parcelle, l.type_local_dvf, l.typologie, l.typologie_confiance,
       l.surface_reelle_bati, l.surface_terrain, l.nb_pieces, l.prix_m2,
       l.dpe_classe, m.id_mutation
FROM dvf_locaux l
JOIN dvf_mutations m ON m.id_mutation = l.id_mutation
WHERE ST_Intersects(m.geom, ST_Transform(ST_GeomFromGeoJSON($1), 2154))
ORDER BY m.date_mutation DESC, m.id_mutation
LIMIT {PLAFOND_LIGNES}
"""

ENTETES = [
    "Date mutation", "Nature", "Valeur foncière (€)", "Commune (INSEE)",
    "Parcelle", "Type de local (DVF)", "Typologie", "Confiance typologie",
    "Surface bâtie (m²)", "Surface terrain (m²)", "Pièces", "Prix (€/m²)",
    "DPE", "Id mutation",
]


async def xlsx_transactions(zone_input: ZoneInput) -> bytes | None:
    """Classeur xlsx des transactions de la zone de contexte, ou None sans base."""
    p = await db.pool()
    if p is None:
        return None
    zone = resolve_zone(zone_input)
    rows = await p.fetch(SQL, json.dumps(zone.large_wgs84.__geo_interface__))
    # openpyxl est synchrone : ne pas bloquer la boucle asyncio pendant l'écriture.
    return await asyncio.to_thread(_construire, [tuple(r) for r in rows], zone.resume)


def _construire(rows: list[tuple], zone_resume: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions DVF"
    ws.append(ENTETES)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(list(row))
    # Formats lisibles sans y passer : montants séparés par milliers, dates ISO courtes.
    for col, fmt in ((1, "yyyy-mm-dd"), (3, "# ##0"), (12, "# ##0")):
        lettre = get_column_letter(col)
        for cell in ws[lettre][1:]:
            cell.number_format = fmt
    largeurs = [12, 10, 16, 10, 16, 18, 22, 10, 14, 15, 8, 11, 6, 24]
    for i, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.auto_filter.ref = ws.dimensions

    # Piste d'audit du fichier : périmètre, volume, source — et alerte si tronqué.
    import datetime

    infos = wb.create_sheet("À propos")
    infos.column_dimensions["A"].width = 26
    infos.column_dimensions["B"].width = 80
    lignes_infos = [
        ("Généré le", datetime.datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Périmètre", "Zone de contexte de l'analyse (grand rayon / polygone)"),
        ("Surface (m²)", zone_resume.get("surface_zone_contexte_m2")),
        ("Transactions exportées", len(rows)),
        ("Source", "DVF géolocalisé (DGFiP / Etalab) — https://files.data.gouv.fr/geo-dvf/"),
        ("Granularité", "Une ligne par local vendu (une mutation peut porter plusieurs locaux)"),
    ]
    if len(rows) >= PLAFOND_LIGNES:
        lignes_infos.insert(
            4,
            ("⚠ Export tronqué", f"Plafond de {PLAFOND_LIGNES} lignes atteint : réduire la zone "
                                 "pour un export exhaustif (les plus récentes sont incluses)."),
        )
    for k, v in lignes_infos:
        infos.append([k, v])
        infos[f"A{infos.max_row}"].font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
