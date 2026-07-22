"""Construction du classeur Excel : structure, formats, alerte de troncature."""
import datetime
import io

from openpyxl import load_workbook

from app import dvf_export

ZONE_RESUME = {"surface_zone_contexte_m2": 123456}


def _ligne(i: int) -> tuple:
    return (
        datetime.date(2025, 1, 1), "Vente", 250000 + i, "69123", f"69123000AB{i:04d}",
        "Appartement", "residentiel", "dvf", "haute", 60, None, 3, 4200, "D", f"2025-{i}",
    )


def _relire(contenu: bytes):
    return load_workbook(io.BytesIO(contenu))


def test_classeur_structure():
    wb = _relire(dvf_export._construire([_ligne(i) for i in range(5)], ZONE_RESUME))
    ws = wb["Transactions DVF"]
    assert [c.value for c in ws[1]] == dvf_export.ENTETES
    assert ws.max_row == 6  # entête + 5 lignes
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None


def test_typologie_en_libelle_metier():
    wb = _relire(dvf_export._construire([_ligne(1)], ZONE_RESUME))
    ligne = [c.value for c in wb["Transactions DVF"][2]]
    assert ligne[dvf_export.COL_TYPOLOGIE] == "Résidentiel"


def test_feuille_a_propos_sans_troncature():
    wb = _relire(dvf_export._construire([_ligne(1)], ZONE_RESUME))
    valeurs = {r[0]: r[1] for r in wb["À propos"].iter_rows(values_only=True)}
    assert valeurs["Transactions exportées"] == 1
    assert "⚠ Export tronqué" not in valeurs


def test_alerte_troncature_au_plafond():
    lignes = [_ligne(i) for i in range(dvf_export.PLAFOND_LIGNES)]
    wb = _relire(dvf_export._construire(lignes, ZONE_RESUME))
    valeurs = {r[0]: r[1] for r in wb["À propos"].iter_rows(values_only=True)}
    assert "⚠ Export tronqué" in valeurs


def test_classeur_vide_reste_valide():
    wb = _relire(dvf_export._construire([], ZONE_RESUME))
    assert wb["Transactions DVF"].max_row == 1
